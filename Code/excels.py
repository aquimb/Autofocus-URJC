import openpyxl as op
import numpy as np
import os
from openpyxl.chart import LineChart, Reference
import math

# This part must be filled and uncomented 
"""

direct_res = Path to the results of the analysis
tissue = ["Adipose/", "Stomach/", "Intestine/", "Kidney/"]
magnification = ["5x/", "10x/"]
depth = ["8 bits/", "16 bits/"]
num_criteria = 4   # Number of criteria evaluated: Accuracy, Range, False maxima, FWHW
num_algs = 15    # Number of algorithms to evaluate

"""


# Names of the algorithms
names = ['Absolute Tenengrad', 'Brener Gradient', 'Entropy', 'First Gaussian Derivative', 'Image Power', 'Laplacian',
         'Normalized Variance', 'Square Gradient', 'Tenengrad', 'Thresholded Absolute Gradient',
         'Thresholded Pixel Count', 'Variance',
         'Variance of log Histogram', 'Vollath4', 'Vollath5']

rank_table = np.array([[0. for a in range(num_criteria)] for b in range(num_algs)])

cwd = os.getcwd()
os.chdir(direct_res)


# Function to convert strings to ints/floats
def str_to_int_or_float(value):
    if isinstance(value, bool):
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value


for tiss in range(4):  # Tissues studied: Adipose, Stomach, Intestine, Kidney (4)
    for mag in range(2):  # Magnifications applied: 5x, 10x (2)
        workbook = op.Workbook()

        for dep in range(2):  # Depths used: 8 bits, 16 bits (2)

            # Creates a summary sheet for each depth
            ws1 = workbook.create_sheet("Sheet_A", 0)
            ws1.title = depth[dep][:-1]
            sheet1 = workbook[depth[dep][:-1]]

            # Writes the names of the seccions in the summary sheets
            sheet1.cell(row=1, column=2, value="Semi-quantitative rankings")
            sheet1.cell(row=19, column=2, value="Semi-quantitative analysis")
            sheet1.merge_cells('B19:H19')
            sheet1.cell(row=20, column=2, value="Total scores")
            sheet1.merge_cells('B20:E20')
            sheet1.cell(row=20, column=7, value="Global score")
            sheet1.cell(row=20, column=8, value="Ranking")

            sheet1.cell(row=39, column=2, value="Quantitative: squared distances")
            sheet1.cell(row=57, column=2, value="Quantitative analysis")
            sheet1.merge_cells('B57:H57')
            sheet1.cell(row=58, column=2, value="Total scores")
            sheet1.merge_cells('B58:E58')
            sheet1.cell(row=58, column=7, value="Global score")
            sheet1.cell(row=58, column=8, value="Ranking")
            for x in range(1, 16):
                sheet1.cell(row=x + 2, column=1, value=names[x - 1]) # Copia datos semi
                sheet1.cell(row=x + 20, column=1, value=names[x - 1]) # Analisis semi
                sheet1.cell(row=x + 40, column=1, value=names[x - 1]) # Copia datos quant
                sheet1.cell(row=x + 58, column=1, value=names[x - 1]) #  Analisis quant

            # Indicates the number of stacks
            if (tiss == 0) and (mag == 1):
                num_stacks = 5
            else:
                num_stacks = 10

            columna = 2
            for k in range(num_stacks):

                # Open the txt file with the data
                path1 = direct_res + tissue[tiss] + magnification[mag] + depth[dep] + "data" + str(k) + ".txt"
                f1 = open(path1, 'r')
                all_lines = f1.readlines()

                # Open the txt file with the criteria tables
                path2 = direct_res + tissue[tiss] + magnification[mag] + depth[dep] + "crit_table" + str(k) + ".txt"
                f2 = open(path2, 'r')
                all_lines2 = f2.readlines()

                # In each tissue and magnification, creates an excel sheet for each depth and stack
                ws1 = workbook.create_sheet("Sheet_A")
                ws1.title = tissue[tiss][0] + " " + magnification[mag][:-1] + " " + depth[dep][:-1] + " S" + str(k)
                sheet = workbook[tissue[tiss][0] + " " + magnification[mag][:-1] + " " + depth[dep][:-1] + " S" + str(k)]

                # Writes the names of the seccions in the depth-stack sheets
                sheet.cell(row=37, column=1, value="Mean")
                sheet.cell(row=38, column=1, value="Std")
                sheet.cell(row=19, column=7, value="Crit. 1")
                sheet.cell(row=19, column=8, value="Crit. 2")
                sheet.cell(row=19, column=9, value="Crit. 3")
                sheet.cell(row=19, column=10, value="Crit. 4")
                sheet.cell(row=18, column=2, value="Criteria Table")
                sheet.merge_cells('B18:E18')
                sheet.cell(row=18, column=7, value="Semi-quantitative ranking")
                sheet.merge_cells('G18:J18')
                sheet.cell(row=40, column=2, value="Normalized values")
                sheet.merge_cells('B40:E40')
                sheet.cell(row=35, column=1, value="Ideal Function")
                sheet.cell(row=56, column=1, value="Ideal Function")
                sheet.cell(row=35, column=2, value=0)
                sheet.cell(row=35, column=3, value=0)
                sheet.cell(row=35, column=4, value=0)
                sheet.cell(row=35, column=5, value=0)
                sheet.cell(row=40, column=7, value="Distance to the ideal value")
                sheet.merge_cells('G40:J40')
                sheet.cell(row=40, column=12, value="Squared distances")
                sheet.merge_cells('L40:O40')
                for x in range(41, 56):
                    sheet.cell(row=x, column=1, value=names[x - 41])

                # Copies to the sheet the data from the txt
                row = 0
                for x in all_lines:
                    one_line = x[:-1].split(';')
                    row_length = len(one_line)

                    for i in range(len(one_line) - 1):
                        one_line[i] = str_to_int_or_float(one_line[i])
                        sheet[chr(i + 65) + str(row + 1)] = one_line[i]
                    row = row + 1

                # Copies to the sheet the criteria tables from the txt
                row = 0
                for x in all_lines2:
                    one_line = x[:-1].split(';')

                    for i in range(len(one_line) - 1):
                        one_line[i] = str_to_int_or_float(one_line[i])
                        sheet[chr(i + 65) + str(row + 19)] = one_line[i]
                    row = row + 1

                # Closes the txt
                f1.close()
                f2.close()

				# Calculates the rest of the data
                for x in range(2, 6):
                    column = [0 for i in range(15)]
                    arr_res = [0 for i in range(15)]

                    for y in sheet.iter_cols(min_row=20, min_col=x, max_row=34, max_col=x, values_only=True):
                        column = y

                        arr = np.array(column)
                        arr_copy = np.array(sorted(set(arr)))
                        offset = 1

                        # Calculates the semi-qualitative rankings for each criterion
                        for j in range(arr_copy.size):
                            indexes = (np.where(arr == (arr_copy[j])))[0]
                            for a in range(indexes.size):
                                arr_res[indexes[a]] = offset
                            offset = offset + indexes.size

                        # And writes the results
                        row = 20
                        for item in arr_res:
                            cell = sheet[chr(x + 64 + 5) + str(row)]
                            cell.value = item
                            row = row + 1

                        # Calculates the mean and the standart deviation, and normalizes the values from the criteria table
                        arr_to_norm = np.append(arr, [0]) # Adds the 0 value for the normalization
                        mean_col = np.mean(arr_to_norm)
                        std_col = np.std(arr_to_norm)
                        sheet[chr(x + 64) + str(37)] = mean_col
                        sheet[chr(x + 64) + str(38)] = std_col

                        if std_col != 0:
                            norm_vals = (arr_to_norm - mean_col) / std_col
                        else:
                            norm_vals = arr_to_norm - mean_col

                        # And writes the results
                        row = 41
                        for item in norm_vals:
                            cell = sheet[chr(x + 64) + str(row)]
                            cell.value = item
                            row = row + 1

                    # Calculates the distances and writes the values
                    for z in range(41, 56):
                        cell_o = sheet[chr(x + 64) + str(z)]
                        cell_d = sheet[chr(x + 64 + 5) + str(z)]
                        cell_r = sheet[chr(x + 64) + str(56)]
                        cell_d.value = cell_r.value - cell_o.value

                    # Calculates the squared distances and writes the values
                    for z in range(41, 56):
                        cell_o = sheet[chr(x + 64 + 5) + str(z)]
                        cell_d = sheet[chr(x + 64 + 10) + str(z)]
                        cell_d.value = cell_o.value * cell_o.value

                # Copies the semi-quantitative ranking table, and the squared distances to the summary sheets
                for x in range(7, 11):
                    # Semi-quantitative
                    for y in range(20, 35):
                        data = sheet[chr(x + 64) + str(y)].value
                        sheet1.cell(row=y - 17, column=columna + x - 7, value=data)

                    # Quantitative
                    for y in range(41, 56):
                        # data = sheet[chr(x + 64) + str(y)].value
                        data = sheet[chr(x + 64 + 5) + str(y)].value  # distancias al cuadrado
                        sheet1.cell(row=y, column=columna + x - 7, value=data)
                columna = columna + 5

                # Creates the chart showing the data
                values = Reference(sheet, min_col=1, min_row=1, max_col=row_length-1, max_row=15)
                chart = LineChart()
                chart.add_data(values, from_rows=True, titles_from_data=True)
                sheet.add_chart(chart, "L18")

			
			# For the summary sheets, calculates the total and global scores and ranking tor the semi-quantitative and quantitative analysis
            valores_q = np.array([0. for n in range(num_stacks)])
            valores_s = np.array([0. for n in range(num_stacks)])
            for y in range(1, 16):
                for x in range(2, 6):
                    for k in range(num_stacks):
                        sheet1.cell(row=2, column=2 + (k * 5), value=("Stack " + str(k)))
                        sheet1.cell(row=40, column=2 + (k * 5), value=("Stack " + str(k)))
                        valores_s[k] = sheet1.cell(row=y + 2, column=x + (k * 5)).value
                        valores_q[k] = sheet1.cell(row=y + 40, column=x + (k * 5)).value

                    result1 = np.sum(valores_s)
                    result2 = math.sqrt(np.sum(valores_q))

                    sheet1.cell(row=y + 20, column=x, value=result1)
                    sheet1.cell(row=y + 58, column=x, value=result2)


                sheet1["G" + str(y + 20)] = "=SUM(B" + str(y + 20) + ":E" + str(y + 20) + ")"
                sheet1["H" + str(y + 20)] = "=RANK(G" + str(y + 20) + ",G21:G35,1)"

                sheet1["G" + str(y + 58)] = "=SUM(B" + str(y + 58) + ":E" + str(y + 58) + ")"
                sheet1["H" + str(y + 58)] = "=RANK(G" + str(y + 58) + ",G59:G73,1)"

        workbook.remove(workbook["Sheet"])
        workbook.save(filename=tissue[tiss][0] + " " + magnification[mag][:-1] + ".xlsx")
